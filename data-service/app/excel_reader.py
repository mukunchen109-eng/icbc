import re
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

SHEET_NAME = "结果"
REQUIRED_COLUMNS = {"id", "text_content", "industry", "area", "title"}


#日期统一化
def parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    match = re.search(
        r"(20\d{2})\D{0,3}(\d{1,2})\D{0,3}(\d{1,2})",
        str(value or ""),
    )
    if not match:
        raise ValueError(f"无法识别日期：{value}")

    year, month, day = map(int, match.groups())
    return date(year, month, day)


def read_batches(
    file_path: str|Path,
    target_date: str|date|None = None,
) -> list[dict]:
    path = Path(file_path)
    if not path.is_file():
        raise FileNotFoundError(f"找不到Excel文件：{path}")

    selected_date = parse_date(target_date) if target_date else None
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"找不到工作表：{SHEET_NAME}")

        sheet = workbook[SHEET_NAME]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)

        if header_row is None:
            raise ValueError("Excel内容为空")

        headers = [
            str(value).strip() if value is not None else ""
            for value in header_row
        ]

        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"Excel缺少字段：{', '.join(sorted(missing))}")

        batches = []

        for row_number, values in enumerate(rows, start=2):
            if not any(value is not None for value in values):
                continue

            row = dict(zip(headers, values))
            batch_date = parse_date(row["title"]) #模拟数据中title列存的日期

            if selected_date and batch_date != selected_date:
                continue

            raw_html = str(row["text_content"] or "").strip()
            if not raw_html:
                raise ValueError(f"第{row_number}行正文为空")

            source_row_id = str(row["id"] or "").strip()
            if not source_row_id:
                raise ValueError(f"第{row_number}行ID为空")

            batches.append(
                {
                    "source_row_id": source_row_id,
                    "news_date": batch_date,
                    "raw_html": raw_html,
                    "industry": row["industry"],
                    "area": row["area"],
                }
            )
        return batches
    finally:
        workbook.close()